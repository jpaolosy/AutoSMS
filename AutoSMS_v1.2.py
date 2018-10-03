""" AutoSMS StopLog """
autosms_version = "1.3" # October 13, 2017 - Slowdown Alert

# IMPORTS
import imapclient
import smtplib
import pyzmail
import openpyxl
import os
import time
import sys
import json
import re

from timeit import default_timer as timer
from datetime import datetime, timedelta


# INITIATE GLOBAL VARIABLES
email='someone@gmail.com'
password='password'
bulk_sms_code='CODE123'
smtp_process_success = "NO"


# CORE FUNCTIONS

def readMail():
    
    global smtp_process_success   
    print("Starting IMAP process...")
    try:     
        imap_obj = imapclient.IMAPClient('imap.gmail.com', ssl=True)
        print("Connected to IMAP")
        updateCounter("imap_success")
        imap_obj.login(email, password)
        imap_obj.select_folder('INBOX', readonly=True)
        print("Checking for new alerts...")
        uids=imap_obj.search('UNSEEN')
        
        if len(uids) < 1:
            print("\n[No new alerts found]")
        else:
            print("\n[" + str(len(uids)) + " new alert/s found!]")
            max_uids = max(uids)
            for uid in uids:
                raw_message = imap_obj.fetch([uid], ['BODY[]', 'FLAGS'])
                message = pyzmail.PyzMessage.factory(raw_message[uid]['BODY[]'])
                mail_subj = message.get_subject()
                message_html = message.html_part.get_payload().decode(message.html_part.charset)
                plant_site, alert_type, alert_code = subjectParser(mail_subj)
                message_body = bodyParser(message_html, alert_type, plant_site)
                mail_to = getMailingList(plant_site, alert_code)
                sendMail(uid, max_uids, plant_site, alert_type, message_body, mail_to)
                saveLastSentAlert(alert_type, plant_site, message_body)
        if smtp_process_success == "YES":
            imap_obj.select_folder('INBOX', readonly=False)
            for uid in uids:
                mark_as_read = imap_obj.fetch([uid], ['BODY[]', 'FLAGS'])
        print("\nDisconnecting from IMAP...")
        imap_obj.logout()
        print("Disconnected from IMAP\n")
        
    except:
        updateCounter("imap_failed")
        imap_resolution = 20
        sleepTime(imap_resolution)
        try:
            print("IMAP process error!")
            imap_obj.logout()
        except:
            print("IMAP connection failed!!!")
   

def sendMail(uid, max_uids, plant_site, alert_type, message_body, mail_to):

    global smtp_process_success
    print("Starting SMTP process...")
    try:
        try:
            print("Trying SMTP SSL")
            smtp_obj = smtplib.SMTP_SSL('smtp.gmail.com', 465)
            #smtp_obj.set_debuglevel(1)
        except:
            print("SMTP SSL connection failed!")
            smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
            #smtp_obj.set_debuglevel(1)
            smtp_obj.ehlo()
            smtp_obj.starttls()
        
        print("Connected to SMTP")
        updateCounter("smtp_success")
        smtp_obj.login(email, password)
        print("Logged on to SMTP")
        
        if uid <= max_uids:
            print("Sending " + alert_type + " message to " + plant_site + " group:")
            for recipients in mail_to:
                print("\t" + recipients)
                smtp_obj.sendmail(email, recipients, "Subject: " + str(bulk_sms_code) + "\n" + str(message_body))
                updateCounter("
                              ")
            print("\nMessage: " + message_body)
            updateCounter("alerts_delivered")
            print("Disconnecting from SMTP...")
            smtp_obj.quit()
            print("Disconnected from SMTP\n")
            sleepTime(5)
        if uid == max_uids:
            smtp_process_success = "YES"
                    
    except:
        updateCounter("smtp_failed")
        smtp_resolution = 30
        sleepTime(smtp_resolution)
        try:
            print("SMTP connection failed!")
            smtp_obj.quit()
        except:
            print("SMTP process error!")
            

def getMailingList(plant_site, alert_code):

    print("Getting mailing list...")
    os.chdir('/home/pi/Desktop/AutoSMS/directory')
    work_book = openpyxl.load_workbook('directory.xlsx', data_only=True)
    lookup_sheet = work_book.get_sheet_by_name('LOOKUP')
    directory_sheet = work_book.get_sheet_by_name(plant_site)

    # PLANT SITE LOOKUP
    for r in range(1,8):
        plant_site_lookup = lookup_sheet.cell(row=r, column=8).value
        if plant_site_lookup == plant_site:
            row_count=lookup_sheet.cell(row=r, column=9).internal_value

    # APPEND CONTACTS TO MAILING LIST      
    mail_to = []
    for r in range (2, row_count+2):
        if directory_sheet['F' + str(r)].internal_value.find(alert_code) > -1:
            mail_to.append(directory_sheet['E' + str(r)].internal_value)
    return mail_to


def bodyParser(message_html, alert_type, plant_site):
	
    plant_site_converted = plantSiteConverter(plant_site)
    
    if message_html != None:

        if alert_type == "StopLog":
            html_len = 31500
            start_len = message_html.find("<BODY>",html_len)+6
            end_len = message_html.find("</BODY>",html_len)
            sliced_body = message_html[start_len:end_len]
            date_end_len = sliced_body.find(":00 ",1)+3
            message_body_prefix = sliced_body[:(date_end_len-19)]
            stop_datetime = datetimeFormatter(alert_type, sliced_body)
            corrected_message_body = message_body_prefix + stop_datetime
            message_body = corrected_message_body

        elif alert_type == "SlowDown":
            html_len = 31500
            start_len = message_html.find("<BODY>",html_len)+6
            end_len = message_html.find("</BODY>",html_len)
            sliced_body = message_html[start_len:end_len]
            date_end_len = sliced_body.find(":00 ",1)+3
            message_body_prefix = sliced_body[:(date_end_len-19)]
            stop_datetime = datetimeFormatter(alert_type, sliced_body)
            stop_datetime = stopTimeSince(0.5, stop_datetime)
            corrected_message_body = message_body_prefix + stop_datetime
            message_body = corrected_message_body
            slowdown_threshold = slowdownThreshold(plant_site)
            slowdown_threshold = ("below " + str(slowdown_threshold) + " tph")
            message_body = message_body.replace('slowdown_threshold', slowdown_threshold)
  
        elif alert_type == "SlowDown_2":
            slowdown_timer, stop_datetime = signalValReportRegEx(message_html, alert_type)
            stop_datetime = stopTimeSince(slowdown_timer, stop_datetime)
            slowdown_threshold = slowdownThreshold (plant_site)
            message_body = ("\n" + plant_site_converted + " kiln is still on SlowDown (below " + str(slowdown_threshold) +" tph) for more than " + str(slowdown_timer) + " hours since " + stop_datetime)

    else:
        print("Mail body is empty!")

    return message_body

    
def subjectParser(mail_subj):
	
    if mail_subj != None:
        plant_site = mail_subj[-4:-1]
        end_len = mail_subj.find(" ")
        alert_type = mail_subj[:end_len]
        
        if alert_type == "StopLog":
            alert_code = "S"
        elif alert_type == "DailyProduction":
            alert_code = "P"
        elif alert_type == "DailyQuality":
            alert_code = "Q"
        elif alert_type == "Environment":
            alert_code = "E"
        elif alert_type == "SlowDown" or alert_type == "SlowDown_2":
            alert_code = "Z"
 
    else:
        print("Mail subject is empty!")

    return plant_site, alert_type, alert_code


def plantSiteConverter(plant_site):
	
    if plant_site == "BTG":
        plant_site_converted = "Batangas"
    elif plant_site == "BUL":
        plant_site_converted = "Bulacan"
    elif plant_site == "NOR":
        plant_site_converted = "Norzagaray"
    elif plant_site == "TER":
        plant_site_converted = "Teresa"
    elif plant_site == "ILG":
        plant_site_converted = "Iligan"
    elif plant_site == "DAN":
        plant_site_converted = "Danao"
    
    return plant_site_converted


def datetimeFormatter(alert_type, message_body):

    message_body = message_body.replace(':', '.')
    
    if alert_type == "StopLog":
        stop_datetime = re.search('\d+.\d+.\d+\s\d+.\d+.\d+',message_body)
        stop_datetime = stop_datetime.group()
    elif alert_type == "SlowDown":
        stop_datetime = re.search('\d+.\d+.\d+\s\d+.\d+.\d+',message_body)
        stop_datetime = stop_datetime.group()
    elif alert_type == "SlowDown_2":
        stop_datetime = message_body
    
    stop_datetime = datetime.strptime(stop_datetime, "%d.%m.%Y %H.%M.%S")
    stop_datetime = stop_datetime.strftime("%m.%d.%Y %I.%M %p")

    return stop_datetime


def stopTimeSince(slowdown_timer, stop_datetime):

    stop_datetime = datetime.strptime(stop_datetime, "%m.%d.%Y %I.%M %p")
    stop_datetime = stop_datetime - (timedelta(hours=slowdown_timer))
    stop_datetime = stop_datetime.strftime("%m.%d.%Y %I.%M %p")
    
    return stop_datetime

	    
def signalValReportRegEx(message_html, alert_type):

    if alert_type == "SlowDown_2":
        html_len = 25000
        start_len = message_html.find("<tr class=\"BMITRodd\">",html_len)
        end_len = (start_len+300)
        sliced_body = message_html[start_len:end_len]
        slowdown_timer = re.search(r'>\d\d\d+<',sliced_body)
        slowdown_timer = slowdown_timer.group()
        slowdown_timer = int(slowdown_timer[1:-1])
        slowdown_timer = round((slowdown_timer/60))
        slowdown_timer = int(slowdown_timer)
        stop_datetime = re.search(r'>\d+.\d+.\d+\s\d+:\d+:\d+<',sliced_body)
        stop_datetime = stop_datetime.group()
        stop_datetime = stop_datetime[1:-1]
        stop_datetime = datetimeFormatter(alert_type, stop_datetime)
        
    else:
        pass

    return slowdown_timer, stop_datetime

def slowdownThreshold (plant_site):

    if plant_site == "BTG":
        slowdown_threshold = 200
    elif plant_site == "BUL":
        slowdown_threshold = 210
    elif plant_site == "NOR":
        slowdown_threshold = 165
    elif plant_site == "TER":
        slowdown_threshold = 200
    elif plant_site == "ILG":
        slowdown_threshold = 200
    elif plant_site == "DAN":
        slowdown_threshold = 200

    return slowdown_threshold


def sleepTime(resolution_time):
    
    time.sleep(resolution_time)


# AUX FUNCTIONS

def updateCounter(counter_object):
        
    with open ("/home/pi/Desktop/AutoSMS/json/counterLog.json") as counter_append:
        json_data = json.load(counter_append)
    if counter_object != "credits_available":
        json_data["counterLog"][counter_object] += 1
    elif counter_object == "credits_available":
        json_data["counterLog"][counter_object] -= 1
    with open("/home/pi/Desktop/AutoSMS/json/counterLog.json", "w") as counter_save:
        counter_save.write(json.dumps(json_data))
        

def printCounterLogValues():
    
    with open ("/home/pi/Desktop/AutoSMS/json/counterLog.json") as counter_values:
        json_data = json.load(counter_values)

    imap_failed = json_data["counterLog"]["imap_failed"]
    imap_success = json_data["counterLog"]["imap_success"]
    smtp_failed = json_data["counterLog"]["smtp_failed"]
    smtp_success = json_data["counterLog"]["smtp_success"]
    alerts_delivered = json_data["counterLog"]["alerts_delivered"]
    credits_available = json_data["counterLog"]["credits_available"]

    print("Sent:" + str(alerts_delivered) + "\t\tCredits:" + str(credits_available) + "\nIMAP:" + str(imap_success) + "/" + str(imap_failed) + "\tSMTP:" + str(smtp_success) + "/" + str(smtp_failed))


def saveLastSentAlert(alert_type, plant_site, message_body):
	
    if alert_type ==  "StopLog":

        if message_body.find("resumed") > 0:
            stoplog_type = "Resume"
        else:
            stoplog_type = "Stop"

        last_sent_alert = (plant_site + " " + stoplog_type + " " + message_body[-19:])

    elif alert_type == "SlowDown" or alert_type == "SlowDown_2":
        last_sent_alert = (plant_site + " " + alert_type + " " + message_body[-19:])


    with open ("/home/pi/Desktop/AutoSMS/json/last_sent.json") as json_load:
        json_data = json.load(json_load)

    json_data["last_sent"]["last_sent"] = last_sent_alert

    with open("/home/pi/Desktop/AutoSMS/json/last_sent.json", "w") as json_save:
        json_save.write(json.dumps(json_data))


def printLastSentAlert():
  
    with open ("/home/pi/Desktop/AutoSMS/json/last_sent.json") as json_load:
        json_data = json.load(json_load)
        
    last_sent_alert = json_data["last_sent"]["last_sent"]
    
    print (last_sent_alert)
    

def serviceStatus(status):
    
    with open ("/home/pi/Desktop/AutoSMS/json/service_status.json") as json_load:
        json_data = json.load(json_load)

    json_data["service_status"]["service_status"] = status
    
    with open("/home/pi/Desktop/AutoSMS/json/service_status.json", "w") as json_save:
        json_save.write(json.dumps(json_data))
        

def setupServiceStart():
    
    with open ("/home/pi/Desktop/AutoSMS/json/setup.json") as setup:
        json_data = json.load(setup)
        
    service_start = json_data["setup_values"]["service_start"]
    
    return service_start


# MAIN FUNCTION

def main():

    while (True):
        if setupServiceStart() == "ON":
            serviceStatus("UP")
            os.system("clear")
            print("\tAutoSMS " + autosms_version)
            print("-" * 30)
            start_timer = timer()
            printCounterLogValues()
            print("-" * 30)
            print(">>>Starting AutoSMS")
            readMail()
            end_timer = timer()
            elapsed_time = round((end_timer-start_timer),2)
            print("Time elapsed: " + str(elapsed_time) + " seconds")
            print(">>>Restarting AutoSMS")
            sleepTime(120) # Program resolution
            os.system("clear")
        elif setup_serviceStart() != "ON":
            serviceStatus("DOWN")
            os.system("clear")
            print("\tAutoSMS " + autosms_version)
            print("-" * 30)
            printCounterValues()
            print("-" * 30)
            print("AutoSMS Service Stopped")
            sleepTime(5)
            os.system("clear")
            

if __name__ == "__main__":

    main()
